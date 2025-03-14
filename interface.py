import tkinter as tk
from tkinter import messagebox, ttk, filedialog, simpledialog
from PIL import Image, ImageTk
import database
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import sys
import os

# Inicializar o banco de dados
database.create_tables()

# Variáveis globais
root = tk.Tk()
root.title("Sistema de Compras")
root.geometry("1400x800")

# Função para determinar o caminho base
def get_base_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    else:
        return os.path.dirname(os.path.abspath(__file__))

# Carregar o logotipo como fundo
def load_background():
    try:
        base_path = get_base_path()
        logo_path = os.path.join(base_path, "Logo.jpg")
        if not os.path.exists(logo_path):
            raise FileNotFoundError("Logo.jpg não encontrado.")
        img = Image.open(logo_path)
        img = img.resize((1400, 800), Image.Resampling.LANCZOS)
        background_image = ImageTk.PhotoImage(img)
        background_label = tk.Label(root, image=background_image)
        background_label.image = background_image  # Mantém referência para evitar garbage collection
        background_label.place(x=0, y=0, relwidth=1, relheight=1)
    except Exception as e:
        messagebox.showwarning("Aviso", f"Erro ao carregar o logotipo: {e}. Continuando sem fundo.")

# Funções da interface
def add_item():
    description = entry_description.get()
    code = entry_code.get()
    brand = entry_brand.get()
    supplier = combo_supplier_entry.get()  # Usa o valor do Combobox
    price = entry_price.get()
    quantity = entry_quantity.get()
    if not description or not code or not supplier or not price or not quantity:
        messagebox.showerror("Erro", "Preencha todos os campos!")
        return
    try:
        quantity = int(quantity)
        if quantity <= 0:
            raise ValueError("A quantidade deve ser maior que zero!")
    except ValueError as e:
        messagebox.showerror("Erro", str(e) if "quantidade" in str(e) else "Quantidade deve ser um número inteiro!")
        return
    status = "A Comprar"
    suppliers_prices = {supplier: float(price)}
    database.insert_item(description, code, brand, status, quantity, suppliers_prices)
    update_item_list()
    clear_entries()

def update_item_list(items=None):
    list_items.delete(0, tk.END)
    if items is None:
        items = database.get_all_items()
    for item in items:
        item_text = f"Desc: {item[1]} | Código: {item[2]} | Marca: {item[3] or 'N/A'} | Status: {item[4]} | Qtd: {item[5]} | Forn/Preço: {', '.join([f'{s}: R${p:.2f}' for s, p in item[6].items()])}"
        list_items.insert(tk.END, item_text)
        if item[4] == "A Comprar":
            list_items.itemconfig(tk.END, {'fg': 'red'})
        elif item[4] == "Comprado":
            list_items.itemconfig(tk.END, {'fg': 'green'})
        elif item[4] == "Parcialmente Comprado":
            list_items.itemconfig(tk.END, {'fg': 'orange'})

def clear_entries():
    entry_description.delete(0, tk.END)
    entry_code.delete(0, tk.END)
    entry_brand.delete(0, tk.END)
    combo_supplier_entry.set('')  # Limpa o Combobox
    entry_price.delete(0, tk.END)
    entry_quantity.delete(0, tk.END)

def remove_item():
    selected = list_items.curselection()
    if not selected:
        messagebox.showerror("Erro", "Selecione um item!")
        return
    item_id = database.get_all_items()[selected[0]][0]
    database.delete_item(item_id)
    update_item_list()

def mark_as_purchased():
    selected = list_items.curselection()
    if not selected:
        messagebox.showerror("Erro", "Selecione um item!")
        return
    item_id = database.get_all_items()[selected[0]][0]
    item = database.get_all_items()[selected[0]]
    database.update_item(item_id, item[1], item[2], item[3], "Comprado", item[5], item[6])
    update_item_list()

def mark_as_partially_purchased():
    selected = list_items.curselection()
    if not selected:
        messagebox.showerror("Erro", "Selecione um item!")
        return
    item_id = database.get_all_items()[selected[0]][0]
    item = database.get_all_items()[selected[0]]
    database.update_item(item_id, item[1], item[2], item[3], "Parcialmente Comprado", item[5], item[6])
    update_item_list()

def filter_items():
    status = combo_status.get()
    supplier = combo_supplier.get().strip()
    if status and supplier:
        items = database.get_items_by_status_and_supplier(status, supplier)
    elif status:
        items = database.get_items_by_status(status)
    elif supplier:
        items = database.get_items_by_supplier(supplier)
    else:
        items = database.get_all_items()
    update_item_list(items)

def generate_excel():
    supplier = entry_export_supplier.get().strip()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Itens"
    ws.append(["Descrição", "Código", "Marca", "Fornecedor", "Preço", "Quantidade", "Status"])
    items = database.get_all_items() if not supplier else database.get_items_by_supplier(supplier)
    for item in items:
        for s, p in item[6].items():
            ws.append([item[1], item[2], item[3] or 'N/A', s, f"R${p:.2f}", item[5], item[4]])
    wb.save("itens.xlsx")
    messagebox.showinfo("Sucesso", "Planilha gerada como 'itens.xlsx'")
def import_excel():
    try:
        base_path = get_base_path()
        excel_path = os.path.join(base_path, "itens_preenchidos.xlsx")
        if not os.path.exists(excel_path):
            excel_path = filedialog.askopenfilename(
                title="Selecione o arquivo Excel",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=base_path
            )
            if not excel_path:
                messagebox.showwarning("Aviso", "Importação cancelada. Nenhum arquivo selecionado.")
                return
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        updated_items = 0  # Contador para itens atualizados
        added_items = 0    # Contador para itens adicionados
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Desempacotar os 6 primeiros valores e ignorar o restante
            description, code, brand, supplier, price, quantity, *rest = row
            print(f"Lendo da planilha: description={description}, code={code}, brand={brand}, supplier={supplier}, price={price}, quantity={quantity}")

            # Normalizar os valores para comparação
            description = str(description).strip().upper() if description else ""
            code = str(code).strip() if code else ""
            brand = str(brand).strip().upper() if brand else "N/A"
            supplier = str(supplier).strip() if supplier else ""
            status = "A Comprar"  # Status padrão para novos itens

            items = database.get_all_items()
            found_match = False
            for item in items:
                # Normalizar os valores do banco para comparação
                item_description = str(item[1]).strip().upper() if item[1] else ""
                item_code = str(item[2]).strip() if item[2] else ""
                item_brand = str(item[3]).strip().upper() if item[3] else "N/A"

                # Comparar descrição, código e marca
                if (item_description == description and 
                    item_code == code and 
                    item_brand == brand):
                    found_match = True
                    suppliers_prices = item[6]
                    if isinstance(price, str):
                        price = price.replace('R$', '').replace(',', '.').strip()
                    price_value = float(price) if price else 0.0
                    print(f"Atualizando item existente (ID={item[0]}): Fornecedor: {supplier}, Novo preço: {price_value}")
                    suppliers_prices[supplier] = price_value
                    database.update_item(item[0], item[1], item[2], item[3], status, quantity, suppliers_prices)
                    updated_items += 1
                    break

            # Se não encontrou correspondência, adicionar novo item
            if not found_match and description and code:
                try:
                    suppliers_prices = {}
                    if isinstance(price, str):
                        price = price.replace('R$', '').replace(',', '.').strip()
                    price_value = float(price) if price else 0.0
                    quantity_value = float(quantity) if quantity else 0.0
                    suppliers_prices[supplier] = price_value
                    item_id = database.add_item(description, code, brand, quantity_value)  # Adiciona e retorna o ID
                    if item_id:
                        database.update_item(item_id, description, code, brand, status, quantity_value, suppliers_prices)
                        print(f"Adicionado novo item (ID={item_id}): {description}, Preço: {price_value}")
                        added_items += 1
                    else:
                        print(f"Falha ao adicionar item: {description}, {code}")
                except Exception as e:
                    print(f"Erro ao adicionar novo item ({description}, {code}): {e}")

        update_item_list()
        if updated_items > 0 or added_items > 0:
            messagebox.showinfo("Sucesso", f"{updated_items} itens atualizados e {added_items} itens adicionados com sucesso!")
        else:
            messagebox.showwarning("Aviso", "Nenhum item foi atualizado ou adicionado. Verifique se os dados da planilha são válidos.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao importar planilha: {e}")
        print(f"Erro detalhado: {e}")

def generate_pdf():
    supplier = entry_pdf_supplier.get().strip()
    company = database.get_company()
    if not company:
        messagebox.showerror("Erro", "Cadastre a empresa primeiro!")
        return
    selected_indices = list_items.curselection()
    if not selected_indices:
        messagebox.showerror("Erro", "Selecione pelo menos um item para gerar o pedido!")
        return
    items = [database.get_all_items()[i] for i in selected_indices]
    order_number = len(database.get_all_items()) + 1
    c = canvas.Canvas(f"pedido_{order_number}.pdf", pagesize=letter)
    c.drawString(100, 750, f"Pedido #{order_number}")
    c.drawString(100, 730, f"Empresa: {company[1]} - CNPJ: {company[2]}")
    c.drawString(100, 710, f"Comprador: {company[3]}")
    y = 690
    for item in items:
        # Informações do item
        item_info = f"I: {item[1]}  Cdg: {item[2]}  M: {item[3] or 'N/A'}  Qtd: {item[5]}"
        # Preços dos fornecedores
        prices_str = "  Prç: " + ", ".join([f"{s}: R${p:.2f}" for s, p in item[6].items() if not supplier or s.strip() == supplier])
        # Combinar tudo em uma única linha
        full_line = item_info + prices_str
        # Desenhar na mesma linha
        c.drawString(100, y, full_line)
        y -= 30  # Próximo item
    c.save()
    messagebox.showinfo("Sucesso", f"Pedido gerado como 'pedido_{order_number}.pdf'")

def register_company():
    name = entry_company_name.get()
    cnpj = entry_company_cnpj.get()
    buyer_name = entry_buyer_name.get()
    if not name or not cnpj or not buyer_name:
        messagebox.showerror("Erro", "Preencha todos os campos da empresa!")
        return
    database.insert_company(name, cnpj, buyer_name)
    messagebox.showinfo("Sucesso", "Empresa cadastrada com sucesso!")
    company_window.destroy()

def register_supplier():
    name = entry_supplier_name.get()
    cnpj = entry_supplier_cnpj.get()
    seller_name = entry_seller_name.get()
    if not name or not cnpj or not seller_name:
        messagebox.showerror("Erro", "Preencha todos os campos do fornecedor!")
        return
    database.insert_supplier(name, cnpj, seller_name)
    messagebox.showinfo("Sucesso", "Fornecedor cadastrado com sucesso!")
    supplier_window.destroy()
    update_supplier_comboboxes()

def update_supplier_comboboxes():
    suppliers = database.get_suppliers()
    combo_supplier_entry['values'] = suppliers  # Atualiza o Combobox de fornecedores no cadastro
    combo_supplier['values'] = [""] + suppliers
    entry_export_supplier['values'] = [""] + suppliers
    entry_pdf_supplier['values'] = [""] + suppliers

def open_company_window():
    global company_window, entry_company_name, entry_company_cnpj, entry_buyer_name
    company_window = tk.Toplevel(root)
    company_window.title("Cadastrar Empresa")
    tk.Label(company_window, text="Nome:").pack()
    entry_company_name = tk.Entry(company_window)
    entry_company_name.pack()
    tk.Label(company_window, text="CNPJ:").pack()
    entry_company_cnpj = tk.Entry(company_window)
    entry_company_cnpj.pack()
    tk.Label(company_window, text="Nome do Comprador:").pack()
    entry_buyer_name = tk.Entry(company_window)
    entry_buyer_name.pack()
    tk.Button(company_window, text="Cadastrar", command=register_company).pack()

def open_supplier_window():
    global supplier_window, entry_supplier_name, entry_supplier_cnpj, entry_seller_name
    supplier_window = tk.Toplevel(root)
    supplier_window.title("Cadastrar Fornecedor")
    tk.Label(supplier_window, text="Nome:").pack()
    entry_supplier_name = tk.Entry(supplier_window)
    entry_supplier_name.pack()
    tk.Label(supplier_window, text="CNPJ:").pack()
    entry_supplier_cnpj = tk.Entry(supplier_window)
    entry_supplier_cnpj.pack()
    tk.Label(supplier_window, text="Nome do Vendedor:").pack()
    entry_seller_name = tk.Entry(supplier_window)
    entry_seller_name.pack()
    tk.Button(supplier_window, text="Cadastrar", command=register_supplier).pack()

def alter_company():
    companies = database.get_all_companies()
    if not companies:
        messagebox.showerror("Erro", "Nenhuma empresa cadastrada!")
        return
    company_id = simpledialog.askinteger("Alterar Empresa", "Digite o ID da empresa a ser alterada:")
    if company_id:
        for company in companies:
            if company[0] == company_id:
                name = simpledialog.askstring("Alterar Empresa", "Novo nome:")
                cnpj = simpledialog.askstring("Alterar Empresa", "Novo CNPJ:")
                buyer_name = simpledialog.askstring("Alterar Empresa", "Novo nome do comprador:")
                if name and cnpj and buyer_name:
                    database.update_company(company_id, name, cnpj, buyer_name)
                    messagebox.showinfo("Sucesso", "Empresa alterada com sucesso!")
                else:
                    messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
                return
        messagebox.showerror("Erro", "Empresa não encontrada!")

def delete_company():
    companies = database.get_all_companies()
    if not companies:
        messagebox.showerror("Erro", "Nenhuma empresa cadastrada!")
        return
    company_id = simpledialog.askinteger("Excluir Empresa", "Digite o ID da empresa a ser excluída:")
    if company_id:
        for company in companies:
            if company[0] == company_id:
                confirm = messagebox.askyesno("Confirmação", "Tem certeza que deseja excluir esta empresa?")
                if confirm:
                    database.delete_company(company_id)
                    messagebox.showinfo("Sucesso", "Empresa excluída com sucesso!")
                return
        messagebox.showerror("Erro", "Empresa não encontrada!")

def alter_supplier():
    suppliers = database.get_all_suppliers()
    if not suppliers:
        messagebox.showerror("Erro", "Nenhum fornecedor cadastrado!")
        return
    supplier_id = simpledialog.askinteger("Alterar Fornecedor", "Digite o ID do fornecedor a ser alterado:")
    if supplier_id:
        for supplier in suppliers:
            if supplier[0] == supplier_id:
                name = simpledialog.askstring("Alterar Fornecedor", "Novo nome:")
                cnpj = simpledialog.askstring("Alterar Fornecedor", "Novo CNPJ:")
                seller_name = simpledialog.askstring("Alterar Fornecedor", "Novo nome do vendedor:")
                if name and cnpj and seller_name:
                    database.update_supplier(supplier_id, name, cnpj, seller_name)
                    messagebox.showinfo("Sucesso", "Fornecedor alterado com sucesso!")
                    update_supplier_comboboxes()
                else:
                    messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
                return
        messagebox.showerror("Erro", "Fornecedor não encontrado!")

def delete_supplier():
    suppliers = database.get_all_suppliers()
    if not suppliers:
        messagebox.showerror("Erro", "Nenhum fornecedor cadastrado!")
        return
    supplier_id = simpledialog.askinteger("Excluir Fornecedor", "Digite o ID do fornecedor a ser excluído:")
    if supplier_id:
        for supplier in suppliers:
            if supplier[0] == supplier_id:
                confirm = messagebox.askyesno("Confirmação", "Tem certeza que deseja excluir este fornecedor?")
                if confirm:
                    database.delete_supplier(supplier_id)
                    messagebox.showinfo("Sucesso", "Fornecedor excluído com sucesso!")
                    update_supplier_comboboxes()
                return
        messagebox.showerror("Erro", "Fornecedor não encontrado!")

# Carregar o fundo ao iniciar
load_background()

# Criar frames para layout
left_frame = tk.Frame(root)
left_frame.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.Y)

right_frame = tk.Frame(root)
right_frame.pack(side=tk.RIGHT, padx=10, pady=10, fill=tk.BOTH, expand=True)

# Inputs no frame esquerdo
tk.Label(left_frame, text="Descrição:", bg="white").pack(pady=5)
entry_description = tk.Entry(left_frame, width=40)
entry_description.pack(pady=5)

tk.Label(left_frame, text="Código:", bg="white").pack(pady=5)
entry_code = tk.Entry(left_frame, width=40)
entry_code.pack(pady=5)

tk.Label(left_frame, text="Marca:", bg="white").pack(pady=5)
entry_brand = tk.Entry(left_frame, width=40)
entry_brand.pack(pady=5)

tk.Label(left_frame, text="Fornecedor:", bg="white").pack(pady=5)
combo_supplier_entry = ttk.Combobox(left_frame, values=database.get_suppliers(), width=37)  # Combobox para fornecedores
combo_supplier_entry.pack(pady=5)

tk.Label(left_frame, text="Preço:", bg="white").pack(pady=5)
entry_price = tk.Entry(left_frame, width=40)
entry_price.pack(pady=5)

tk.Label(left_frame, text="Quantidade:", bg="white").pack(pady=5)
entry_quantity = tk.Entry(left_frame, width=40)
entry_quantity.pack(pady=5)

# Botões no frame esquerdo
tk.Button(left_frame, text="Adicionar Item", command=add_item).pack(pady=5)
tk.Button(left_frame, text="Remover Item", command=remove_item).pack(pady=5)
tk.Button(left_frame, text="Marcar como Comprado", command=mark_as_purchased).pack(pady=5)
tk.Button(left_frame, text="Marcar como Parcialmente Comprado", command=mark_as_partially_purchased).pack(pady=5)

# Lista de itens no centro com seleção múltipla
list_items = tk.Listbox(root, height=20, width=90, selectmode=tk.MULTIPLE)
list_items.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.BOTH, expand=True)
update_item_list()

# Filtros no frame direito
tk.Label(right_frame, text="Filtrar por Status:", bg="white").pack(pady=5, anchor="w")
combo_status = ttk.Combobox(right_frame, values=["", "A Comprar", "Comprado", "Parcialmente Comprado"], width=20)
combo_status.pack(pady=5, fill=tk.X)
combo_status.set("")

tk.Label(right_frame, text="Filtrar por Fornecedor:", bg="white").pack(pady=5, anchor="w")
combo_supplier = ttk.Combobox(right_frame, values=[""] + database.get_suppliers(), width=40)
combo_supplier.pack(pady=5, fill=tk.X)
combo_supplier.set("")

tk.Button(right_frame, text="Filtrar", command=filter_items, width=20).pack(pady=10)

# Campos para exportar, importar e PDF com fornecedor
tk.Label(right_frame, text="Fornecedor para Exportar Excel:", bg="white").pack(pady=5, anchor="w")
entry_export_supplier = ttk.Combobox(right_frame, values=[""] + database.get_suppliers(), width=40)
entry_export_supplier.pack(pady=5, fill=tk.X)
entry_export_supplier.set("")

tk.Label(right_frame, text="Fornecedor para Pedido PDF:", bg="white").pack(pady=5, anchor="w")
entry_pdf_supplier = ttk.Combobox(right_frame, values=[""] + database.get_suppliers(), width=40)
entry_pdf_supplier.pack(pady=5, fill=tk.X)
entry_pdf_supplier.set("")

tk.Button(right_frame, text="Exportar Excel", command=generate_excel, width=20).pack(pady=5)
tk.Button(right_frame, text="Importar Excel", command=import_excel, width=20).pack(pady=5)
tk.Button(right_frame, text="Gerar Pedido em PDF", command=generate_pdf, width=20).pack(pady=5)

# Menu
menu = tk.Menu(root)
root.config(menu=menu)

cadastro_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="Cadastro", menu=cadastro_menu)
cadastro_menu.add_command(label="Cadastrar Empresa", command=open_company_window)
cadastro_menu.add_command(label="Alterar Empresa", command=alter_company)
cadastro_menu.add_command(label="Excluir Empresa", command=delete_company)
cadastro_menu.add_command(label="Cadastrar Fornecedor", command=open_supplier_window)
cadastro_menu.add_command(label="Alterar Fornecedor", command=alter_supplier)
cadastro_menu.add_command(label="Excluir Fornecedor", command=delete_supplier)

compras_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="Compras", menu=compras_menu)
compras_menu.add_command(label="Gerar Planilha Excel", command=generate_excel)
compras_menu.add_command(label="Importar Planilha Preenchida", command=import_excel)
compras_menu.add_command(label="Gerar Pedido em PDF", command=generate_pdf)

# Inicializar os fornecedores
update_supplier_comboboxes()

root.mainloop()