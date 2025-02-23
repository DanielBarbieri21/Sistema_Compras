import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from PIL import Image, ImageTk
import database
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import sys
import os

# Inicializar o banco de dados
database.create_tables()

# Variáveis globais
root = tk.Tk()
root.title("Sistema de Compras")
root.geometry("1400x800")

# Função para determinar o caminho base
def get_base_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    else:
        return os.path.dirname(os.path.abspath(__file__))

# Carregar o logotipo como fundo
def load_background():
    try:
        base_path = get_base_path()
        logo_path = os.path.join(base_path, "Logo.jpg")
        if not os.path.exists(logo_path):
            raise FileNotFoundError("Logo.jpg não encontrado.")
        img = Image.open(logo_path)
        img = img.resize((1400, 800), Image.Resampling.LANCZOS)
        background_image = ImageTk.PhotoImage(img)
        background_label = tk.Label(root, image=background_image)
        background_label.image = background_image  # Mantém referência para evitar garbage collection
        background_label.place(x=0, y=0, relwidth=1, relheight=1)
    except Exception as e:
        messagebox.showwarning("Aviso", f"Erro ao carregar o logotipo: {e}. Continuando sem fundo.")

# Funções da interface
def add_item():
    description = entry_description.get()
    code = entry_code.get()
    brand = entry_brand.get()
    supplier = entry_supplier.get()
    price = entry_price.get()
    if not description or not code or not supplier or not price:
        messagebox.showerror("Erro", "Preencha todos os campos!")
        return
    status = "A Comprar"
    suppliers_prices = {supplier: float(price)}
    database.insert_item(description, code, brand, status, suppliers_prices)
    update_item_list()
    clear_entries()

def update_item_list(items=None):
    list_items.delete(0, tk.END)
    if items is None:
        items = database.get_all_items()
    for item in items:
        item_text = f"Desc: {item[1]} | Código: {item[2]} | Marca: {item[3] or 'N/A'} | Status: {item[4]} | Fornecedor/Preço: {', '.join([f'{s}: R${p:.2f}' for s, p in item[5].items()])}"
        list_items.insert(tk.END, item_text)
        if item[4] == "A Comprar":
            list_items.itemconfig(tk.END, {'fg': 'red'})
        elif item[4] == "Comprado":
            list_items.itemconfig(tk.END, {'fg': 'green'})
        elif item[4] == "Parcialmente Comprado":
            list_items.itemconfig(tk.END, {'fg': 'orange'})

def clear_entries():
    entry_description.delete(0, tk.END)
    entry_code.delete(0, tk.END)
    entry_brand.delete(0, tk.END)
    entry_supplier.delete(0, tk.END)
    entry_price.delete(0, tk.END)

def remove_item():
    selected = list_items.curselection()
    if not selected:
        messagebox.showerror("Erro", "Selecione um item!")
        return
    item_id = database.get_all_items()[selected[0]][0]
    database.delete_item(item_id)
    update_item_list()

def mark_as_purchased():
    selected = list_items.curselection()
    if not selected:
        messagebox.showerror("Erro", "Selecione um item!")
        return
    item_id = database.get_all_items()[selected[0]][0]
    item = database.get_all_items()[selected[0]]
    database.update_item(item_id, item[1], item[2], item[3], "Comprado", item[5])
    update_item_list()

def mark_as_partially_purchased():
    selected = list_items.curselection()
    if not selected:
        messagebox.showerror("Erro", "Selecione um item!")
        return
    item_id = database.get_all_items()[selected[0]][0]
    item = database.get_all_items()[selected[0]]
    database.update_item(item_id, item[1], item[2], item[3], "Parcialmente Comprado", item[5])
    update_item_list()

def filter_items():
    status = combo_status.get()
    supplier = combo_supplier.get()
    if status and not supplier:
        items = database.get_items_by_status(status)
    elif supplier and not status:
        items = database.get_items_by_supplier(supplier)
    else:
        items = database.get_all_items()
    update_item_list(items)

def generate_excel():
    supplier = entry_export_supplier.get()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Itens"
    ws.append(["Descrição", "Código", "Marca", "Fornecedor", "Preço", "Status"])
    items = database.get_all_items() if not supplier else database.get_items_by_supplier(supplier)
    for item in items:
        for s, p in item[5].items():
            ws.append([item[1], item[2], item[3] or 'N/A', s, f"R${p:.2f}", item[4]])
    wb.save("itens.xlsx")
    messagebox.showinfo("Sucesso", "Planilha gerada como 'itens.xlsx'")

def import_excel():
    try:
        base_path = get_base_path()
        excel_path = os.path.join(base_path, "itens_preenchidos.xlsx")
        if not os.path.exists(excel_path):
            excel_path = filedialog.askopenfilename(
                title="Selecione o arquivo Excel",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=base_path
            )
            if not excel_path:
                messagebox.showwarning("Aviso", "Importação cancelada. Nenhum arquivo selecionado.")
                return
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            description, code, brand, supplier, price, _ = row
            items = database.get_all_items()
            for item in items:
                if item[1] == description and item[2] == code and (item[3] or 'N/A') == (brand or 'N/A'):
                    suppliers_prices = item[5]
                    if isinstance(price, str):
                        price = price.replace('R$', '').replace('.', '').replace(',', '.')
                    suppliers_prices[supplier] = float(price) if price else 0
                    database.update_item(item[0], item[1], item[2], item[3], item[4], suppliers_prices)
        update_item_list()
        messagebox.showinfo("Sucesso", "Preços atualizados com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao importar planilha: {e}")

def generate_pdf():
    supplier = entry_pdf_supplier.get()
    company = database.get_company()
    if not company:
        messagebox.showerror("Erro", "Cadastre a empresa primeiro!")
        return
    selected = list_items.curselection()
    if not selected:
        messagebox.showerror("Erro", "Selecione um item para gerar o pedido!")
        return
    item = database.get_all_items()[selected[0]]
    order_number = len(database.get_all_items()) + 1
    c = canvas.Canvas(f"pedido_{order_number}.pdf", pagesize=letter)
    c.drawString(100, 750, f"Pedido #{order_number}")
    c.drawString(100, 730, f"Empresa: {company[1]} - CNPJ: {company[2]}")
    c.drawString(100, 710, f"Comprador: {company[3]}")
    c.drawString(100, 690, f"Item: {item[1]} - Código: {item[2]} - Marca: {item[3] or 'N/A'}")
    c.drawString(100, 670, f"Status: {item[4]}")
    c.drawString(100, 650, f"Fornecedor: {supplier if supplier else 'Todos'}")
    c.drawString(100, 630, "Preços:")
    y = 610
    for s, p in item[5].items():
        if not supplier or s == supplier:
            c.drawString(100, y, f"{s}: R${p:.2f}")
            y -= 20
    c.save()
    messagebox.showinfo("Sucesso", f"Pedido gerado como 'pedido_{order_number}.pdf'")

def register_company():
    name = entry_company_name.get()
    cnpj = entry_company_cnpj.get()
    buyer_name = entry_buyer_name.get()
    if not name or not cnpj or not buyer_name:
        messagebox.showerror("Erro", "Preencha todos os campos da empresa!")
        return
    database.insert_company(name, cnpj, buyer_name)
    messagebox.showinfo("Sucesso", "Empresa cadastrada com sucesso!")
    company_window.destroy()

def register_supplier():
    name = entry_supplier_name.get()
    cnpj = entry_supplier_cnpj.get()
    seller_name = entry_seller_name.get()
    if not name or not cnpj or not seller_name:
        messagebox.showerror("Erro", "Preencha todos os campos do fornecedor!")
        return
    database.insert_supplier(name, cnpj, seller_name)
    messagebox.showinfo("Sucesso", "Fornecedor cadastrado com sucesso!")
    supplier_window.destroy()
    combo_supplier['values'] = database.get_suppliers()
    entry_export_supplier['values'] = database.get_suppliers()
    entry_pdf_supplier['values'] = database.get_suppliers()

def open_company_window():
    global company_window, entry_company_name, entry_company_cnpj, entry_buyer_name
    company_window = tk.Toplevel(root)
    company_window.title("Cadastrar Empresa")
    tk.Label(company_window, text="Nome:").pack()
    entry_company_name = tk.Entry(company_window)
    entry_company_name.pack()
    tk.Label(company_window, text="CNPJ:").pack()
    entry_company_cnpj = tk.Entry(company_window)
    entry_company_cnpj.pack()
    tk.Label(company_window, text="Nome do Comprador:").pack()
    entry_buyer_name = tk.Entry(company_window)
    entry_buyer_name.pack()
    tk.Button(company_window, text="Cadastrar", command=register_company).pack()

def open_supplier_window():
    global supplier_window, entry_supplier_name, entry_supplier_cnpj, entry_seller_name
    supplier_window = tk.Toplevel(root)
    supplier_window.title("Cadastrar Fornecedor")
    tk.Label(supplier_window, text="Nome:").pack()
    entry_supplier_name = tk.Entry(supplier_window)
    entry_supplier_name.pack()
    tk.Label(supplier_window, text="CNPJ:").pack()
    entry_supplier_cnpj = tk.Entry(supplier_window)
    entry_supplier_cnpj.pack()
    tk.Label(supplier_window, text="Nome do Vendedor:").pack()
    entry_seller_name = tk.Entry(supplier_window)
    entry_seller_name.pack()
    tk.Button(supplier_window, text="Cadastrar", command=register_supplier).pack()

# Carregar o fundo ao iniciar
load_background()

# Criar frames para layout
left_frame = tk.Frame(root)
left_frame.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.Y)

right_frame = tk.Frame(root)
right_frame.pack(side=tk.RIGHT, padx=10, pady=10, fill=tk.BOTH, expand=True)

# Inputs no frame esquerdo
tk.Label(left_frame, text="Descrição:", bg="white").pack(pady=5)
entry_description = tk.Entry(left_frame, width=40)
entry_description.pack(pady=5)

tk.Label(left_frame, text="Código:", bg="white").pack(pady=5)
entry_code = tk.Entry(left_frame, width=40)
entry_code.pack(pady=5)

tk.Label(left_frame, text="Marca:", bg="white").pack(pady=5)
entry_brand = tk.Entry(left_frame, width=40)
entry_brand.pack(pady=5)

tk.Label(left_frame, text="Fornecedor:", bg="white").pack(pady=5)
entry_supplier = tk.Entry(left_frame, width=40)
entry_supplier.pack(pady=5)

tk.Label(left_frame, text="Preço:", bg="white").pack(pady=5)
entry_price = tk.Entry(left_frame, width=40)
entry_price.pack(pady=5)

# Botões no frame esquerdo
tk.Button(left_frame, text="Adicionar Item", command=add_item).pack(pady=5)
tk.Button(left_frame, text="Remover Item", command=remove_item).pack(pady=5)
tk.Button(left_frame, text="Marcar como Comprado", command=mark_as_purchased).pack(pady=5)
tk.Button(left_frame, text="Marcar como Parcialmente Comprado", command=mark_as_partially_purchased).pack(pady=5)

# Lista de itens no centro
list_items = tk.Listbox(root, height=20, width=90)
list_items.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.BOTH, expand=True)

update_item_list()

# Filtros no frame direito
tk.Label(right_frame, text="Filtrar por Status:", bg="white").pack(pady=5, anchor="w")
combo_status = ttk.Combobox(right_frame, values=["A Comprar", "Comprado", "Parcialmente Comprado"], width=40)
combo_status.pack(pady=5, fill=tk.X)

tk.Label(right_frame, text="Filtrar por Fornecedor:", bg="white").pack(pady=5, anchor="w")
combo_supplier = ttk.Combobox(right_frame, values=database.get_suppliers(), width=40)
combo_supplier.pack(pady=5, fill=tk.X)

tk.Button(right_frame, text="Filtrar", command=filter_items, width=20).pack(pady=10)

# Campos para exportar, importar e PDF com fornecedor
tk.Label(right_frame, text="Fornecedor para Exportar Excel:", bg="white").pack(pady=5, anchor="w")
entry_export_supplier = ttk.Combobox(right_frame, values=database.get_suppliers(), width=40)
entry_export_supplier.pack(pady=5, fill=tk.X)

tk.Label(right_frame, text="Fornecedor para Pedido PDF:", bg="white").pack(pady=5, anchor="w")
entry_pdf_supplier = ttk.Combobox(right_frame, values=database.get_suppliers(), width=40)
entry_pdf_supplier.pack(pady=5, fill=tk.X)

tk.Button(right_frame, text="Exportar Excel", command=generate_excel, width=20).pack(pady=5)
tk.Button(right_frame, text="Importar Excel", command=import_excel, width=20).pack(pady=5)
tk.Button(right_frame, text="Gerar Pedido em PDF", command=generate_pdf, width=20).pack(pady=5)

# Menu
menu = tk.Menu(root)
root.config(menu=menu)

cadastro_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="Cadastro", menu=cadastro_menu)
cadastro_menu.add_command(label="Cadastrar Empresa", command=open_company_window)
cadastro_menu.add_command(label="Cadastrar Fornecedor", command=open_supplier_window)

compras_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="Compras", menu=compras_menu)
compras_menu.add_command(label="Gerar Planilha Excel", command=generate_excel)
compras_menu.add_command(label="Importar Planilha Preenchida", command=import_excel)
compras_menu.add_command(label="Gerar Pedido em PDF", command=generate_pdf)

root.mainloop()