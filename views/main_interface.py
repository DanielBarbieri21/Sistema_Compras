"""
Interface Principal Moderna do Sistema de Compras
"""
import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *

# PIL é opcional
try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None
import sys
import os

from models import Item, ItemRepository, Company, CompanyRepository, Supplier, SupplierRepository
from utils import ItemValidator, CompanyValidator, SupplierValidator, ExcelValidator
from services import BackupService
from views.dashboard import DashboardView


class MainInterface:
    """Interface principal moderna do sistema"""
    
    def __init__(self):
        self.root = ttk.Window(themename="superhero")
        self.root.title("Sistema de Compras v2.0")
        self.root.geometry("1400x800")
        
        # Serviços
        self.backup_service = BackupService()
        self.backup_service.start_automatic_backup()
        
        # Variáveis
        self.current_view = None
        
        # Inicializar interface
        self.create_menu()
        self.create_main_interface()
        self.load_suppliers()
        
        # Carregar dashboard por padrão
        self.show_dashboard()
    
    def create_menu(self):
        """Cria o menu principal"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # Menu Sistema
        system_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Sistema", menu=system_menu)
        system_menu.add_command(label="Dashboard", command=self.show_dashboard)
        system_menu.add_separator()
        system_menu.add_command(label="Backup Manual", command=self.create_manual_backup)
        system_menu.add_command(label="Gerenciar Backups", command=self.show_backup_manager)
        system_menu.add_separator()
        system_menu.add_command(label="Sair", command=self.root.quit)
        
        # Menu Cadastros
        cadastro_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Cadastros", menu=cadastro_menu)
        cadastro_menu.add_command(label="Itens", command=self.show_items_view)
        cadastro_menu.add_command(label="Empresas", command=self.show_companies_view)
        cadastro_menu.add_command(label="Fornecedores", command=self.show_suppliers_view)
        
        # Menu Compras
        compras_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Compras", menu=compras_menu)
        compras_menu.add_command(label="Gerar Excel", command=self.generate_excel)
        compras_menu.add_command(label="Importar Excel", command=self.import_excel)
        compras_menu.add_command(label="Gerar PDF", command=self.generate_pdf)
        
        # Menu Ajuda
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ajuda", menu=help_menu)
        help_menu.add_command(label="Sobre", command=self.show_about)
    
    def create_main_interface(self):
        """Cria a interface principal"""
        # Frame principal
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Barra de status
        self.create_status_bar()
    
    def create_status_bar(self):
        """Cria a barra de status"""
        self.status_frame = ttk.Frame(self.root)
        self.status_frame.pack(side=BOTTOM, fill=X, padx=10, pady=5)
        
        self.status_label = ttk.Label(self.status_frame, text="Sistema iniciado")
        self.status_label.pack(side=LEFT)
        
        # Indicador de backup
        self.backup_indicator = ttk.Label(self.status_frame, text="🔄 Backup: Ativo")
        self.backup_indicator.pack(side=RIGHT)
    
    def show_dashboard(self):
        """Mostra o dashboard"""
        self.clear_current_view()
        self.current_view = DashboardView(self.main_frame)
        self.current_view.pack(fill=BOTH, expand=True)
        self.update_status("Dashboard carregado")
    
    def show_items_view(self):
        """Mostra a view de itens"""
        self.clear_current_view()
        self.current_view = ItemsView(self.main_frame, self)
        self.current_view.pack(fill=BOTH, expand=True)
        self.update_status("Gerenciamento de itens")
    
    def show_companies_view(self):
        """Mostra a view de empresas"""
        self.clear_current_view()
        self.current_view = CompaniesView(self.main_frame, self)
        self.current_view.pack(fill=BOTH, expand=True)
        self.update_status("Gerenciamento de empresas")
    
    def show_suppliers_view(self):
        """Mostra a view de fornecedores"""
        self.clear_current_view()
        self.current_view = SuppliersView(self.main_frame, self)
        self.current_view.pack(fill=BOTH, expand=True)
        self.update_status("Gerenciamento de fornecedores")
    
    def clear_current_view(self):
        """Limpa a view atual"""
        if self.current_view:
            self.current_view.destroy()
            self.current_view = None
    
    def load_suppliers(self):
        """Carrega a lista de fornecedores"""
        self.suppliers = SupplierRepository.get_names()
    
    def update_status(self, message):
        """Atualiza a barra de status"""
        self.status_label.config(text=message)
    
    def create_manual_backup(self):
        """Cria backup manual"""
        try:
            backup_path = self.backup_service.create_backup("manual")
            if backup_path:
                messagebox.showinfo("Sucesso", f"Backup criado com sucesso!\n{backup_path}")
            else:
                messagebox.showerror("Erro", "Falha ao criar backup")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar backup: {e}")
    
    def show_backup_manager(self):
        """Mostra o gerenciador de backups"""
        try:
            backups = self.backup_service.list_backups()
            
            # Criar janela de backups
            backup_window = ttk.Toplevel(self.root)
            backup_window.title("Gerenciar Backups")
            backup_window.geometry("700x500")
            
            # Lista de backups
            tree = ttk.Treeview(backup_window, columns=('size', 'created'), show='tree headings')
            tree.heading('#0', text='Arquivo')
            tree.heading('size', text='Tamanho')
            tree.heading('created', text='Criado em')
            
            tree.column('#0', width=400)
            tree.column('size', width=100)
            tree.column('created', width=150)
            
            for backup in backups:
                size_mb = backup['size'] / (1024 * 1024)
                tree.insert('', 'end', text=backup['filename'], 
                           values=(f"{size_mb:.2f} MB", backup['created'].strftime('%d/%m/%Y %H:%M')))
            
            tree.pack(fill=BOTH, expand=True, padx=10, pady=10)
            
            # Botões
            buttons_frame = ttk.Frame(backup_window)
            buttons_frame.pack(fill=X, padx=10, pady=5)
            
            ttk.Button(buttons_frame, text="Restaurar", 
                      command=lambda: self.restore_backup(tree, backup_window)).pack(side=LEFT, padx=5)
            ttk.Button(buttons_frame, text="Excluir", 
                      command=lambda: self.delete_backup(tree, backup_window)).pack(side=LEFT, padx=5)
            ttk.Button(buttons_frame, text="Fechar", 
                      command=backup_window.destroy).pack(side=RIGHT, padx=5)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao mostrar backups: {e}")
    
    def restore_backup(self, tree, window):
        """Restaura um backup selecionado"""
        try:
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("Aviso", "Selecione um backup para restaurar")
                return
            
            item = tree.item(selection[0])
            filename = item['text']
            
            # Encontrar o caminho completo do backup
            backups = self.backup_service.list_backups()
            backup_path = None
            for backup in backups:
                if backup['filename'] == filename:
                    backup_path = backup['path']
                    break
            
            if not backup_path:
                messagebox.showerror("Erro", "Backup não encontrado")
                return
            
            # Confirmar restauração
            if messagebox.askyesno("Confirmar", f"Restaurar backup {filename}?\n\nEsta ação substituirá o banco atual."):
                if self.backup_service.restore_backup(backup_path):
                    messagebox.showinfo("Sucesso", "Backup restaurado com sucesso!")
                    window.destroy()
                    self.load_suppliers()  # Recarregar dados
                else:
                    messagebox.showerror("Erro", "Falha ao restaurar backup")
                    
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao restaurar backup: {e}")
    
    def delete_backup(self, tree, window):
        """Exclui um backup selecionado"""
        try:
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("Aviso", "Selecione um backup para excluir")
                return
            
            item = tree.item(selection[0])
            filename = item['text']
            
            # Encontrar o caminho completo do backup
            backups = self.backup_service.list_backups()
            backup_path = None
            for backup in backups:
                if backup['filename'] == filename:
                    backup_path = backup['path']
                    break
            
            if not backup_path:
                messagebox.showerror("Erro", "Backup não encontrado")
                return
            
            # Confirmar exclusão
            if messagebox.askyesno("Confirmar", f"Excluir backup {filename}?"):
                if self.backup_service.delete_backup(backup_path):
                    messagebox.showinfo("Sucesso", "Backup excluído com sucesso!")
                    window.destroy()
                    self.show_backup_manager()  # Atualizar lista
                else:
                    messagebox.showerror("Erro", "Falha ao excluir backup")
                    
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir backup: {e}")
    
    def generate_excel(self):
        """Gera planilha Excel"""
        try:
            import openpyxl
            
            # Janela para selecionar fornecedor
            supplier_window = ttk.Toplevel(self.root)
            supplier_window.title("Exportar Excel")
            supplier_window.geometry("400x200")
            
            ttk.Label(supplier_window, text="Fornecedor (deixe vazio para todos):").pack(pady=10)
            
            supplier_combo = ttk.Combobox(supplier_window, values=[""] + self.suppliers)
            supplier_combo.pack(pady=5)
            supplier_combo.set("")
            
            def export_excel():
                supplier = supplier_combo.get().strip()
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Itens"
                ws.append(["Descrição", "Código", "Marca", "Fornecedor", "Preço", "Quantidade", "Status"])
                
                items = ItemRepository.get_all()
                for item in items:
                    if supplier:
                        # Filtrar por fornecedor
                        if supplier in item.suppliers_prices:
                            ws.append([
                                item.description, item.code, item.brand, supplier,
                                f"R${item.suppliers_prices[supplier]:.2f}", 
                                item.quantity, item.status
                            ])
                    else:
                        # Todos os fornecedores
                        for s, p in item.suppliers_prices.items():
                            ws.append([
                                item.description, item.code, item.brand, s,
                                f"R${p:.2f}", item.quantity, item.status
                            ])
                
                wb.save("itens.xlsx")
                messagebox.showinfo("Sucesso", "Planilha gerada como 'itens.xlsx'")
                supplier_window.destroy()
            
            ttk.Button(supplier_window, text="Exportar", command=export_excel).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar Excel: {e}")
    
    def import_excel(self):
        """Importa planilha Excel"""
        try:
            file_path = filedialog.askopenfilename(
                title="Selecione o arquivo Excel",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not file_path:
                return
            
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            updated_items = 0
            added_items = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 6:
                    continue
                
                description, code, brand, supplier, price, quantity, *rest = row
                
                # Normalizar dados
                description = str(description).strip().upper() if description else ""
                code = str(code).strip() if code else ""
                brand = str(brand).strip().upper() if brand else "N/A"
                supplier = str(supplier).strip() if supplier else ""
                
                if not description or not code or not supplier:
                    continue
                
                # Validar dados
                errors = ItemValidator.validate_item_data(
                    description, code, brand, supplier, str(price), str(quantity)
                )
                
                if errors:
                    print(f"Erro na linha: {errors}")
                    continue
                
                # Processar preço
                if isinstance(price, str):
                    price = price.replace('R$', '').replace(',', '.').strip()
                price_value = float(price) if price else 0.0
                quantity_value = float(quantity) if quantity else 0.0
                
                # Buscar item existente
                items = ItemRepository.get_all()
                found_match = False
                
                for item in items:
                    if (item.description.upper() == description and 
                        item.code == code and 
                        item.brand.upper() == brand):
                        
                        # Atualizar item existente
                        item.suppliers_prices[supplier] = price_value
                        item.quantity = quantity_value
                        ItemRepository.update(item)
                        updated_items += 1
                        found_match = True
                        break
                
                # Adicionar novo item se não encontrado
                if not found_match:
                    new_item = Item(
                        description=description,
                        code=code,
                        brand=brand,
                        status="A Comprar",
                        quantity=quantity_value,
                        suppliers_prices={supplier: price_value}
                    )
                    ItemRepository.create(new_item)
                    added_items += 1
            
            messagebox.showinfo("Sucesso", 
                              f"{updated_items} itens atualizados e {added_items} itens adicionados!")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao importar Excel: {e}")
    
    def generate_pdf(self):
        """Gera PDF de pedido"""
        try:
            # Import opcional com fallback amigável
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.pdfgen import canvas
            except Exception:
                messagebox.showerror("Dependência ausente", "Biblioteca reportlab não instalada. Instale para gerar PDF.")
                return
            
            # Verificar se há empresa cadastrada
            company = CompanyRepository.get_default()
            if not company:
                messagebox.showerror("Erro", "Cadastre uma empresa primeiro!")
                return
            
            # Janela para selecionar fornecedor
            supplier_window = ttk.Toplevel(self.root)
            supplier_window.title("Gerar PDF")
            supplier_window.geometry("400x200")
            
            ttk.Label(supplier_window, text="Fornecedor (deixe vazio para todos):").pack(pady=10)
            
            supplier_combo = ttk.Combobox(supplier_window, values=[""] + self.suppliers)
            supplier_combo.pack(pady=5)
            supplier_combo.set("")
            
            def generate_pdf_file():
                supplier = supplier_combo.get().strip()
                
                items = ItemRepository.get_all()
                order_number = len(items) + 1
                
                c = canvas.Canvas(f"pedido_{order_number}.pdf", pagesize=letter)
                c.drawString(100, 750, f"Pedido #{order_number}")
                c.drawString(100, 730, f"Empresa: {company.name} - CNPJ: {company.cnpj}")
                c.drawString(100, 710, f"Comprador: {company.buyer_name}")
                
                y = 690
                for item in items:
                    if supplier and supplier not in item.suppliers_prices:
                        continue
                    
                    item_info = f"I: {item.description}  Cdg: {item.code}  M: {item.brand}  Qtd: {item.quantity}"
                    
                    if supplier:
                        price = item.suppliers_prices.get(supplier, 0)
                        prices_str = f"  Prç: {supplier}: R${price:.2f}"
                    else:
                        prices_str = "  Prç: " + ", ".join([f"{s}: R${p:.2f}" for s, p in item.suppliers_prices.items()])
                    
                    full_line = item_info + prices_str
                    c.drawString(100, y, full_line)
                    y -= 30
                
                c.save()
                messagebox.showinfo("Sucesso", f"Pedido gerado como 'pedido_{order_number}.pdf'")
                supplier_window.destroy()
            
            ttk.Button(supplier_window, text="Gerar PDF", command=generate_pdf_file).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar PDF: {e}")
    
    def show_about(self):
        """Mostra informações sobre o sistema"""
        about_text = """
Sistema de Compras v2.0

Desenvolvido com Python e Tkinter
Interface moderna com ttkbootstrap

Funcionalidades:
• Gerenciamento de itens
• Cadastro de empresas e fornecedores
• Dashboard com estatísticas
• Backup automático
• Exportação para Excel e PDF
• Importação de planilhas
• Validação robusta de dados

© 2024 - Sistema de Compras
        """
        
        messagebox.showinfo("Sobre", about_text)
    
    def run(self):
        """Executa a aplicação"""
        self.root.mainloop()
    
    def __del__(self):
        """Destrutor - para o backup automático"""
        if hasattr(self, 'backup_service'):
            self.backup_service.stop_automatic_backup()


# Views específicas serão implementadas em arquivos separados
class ItemsView:
    """View para gerenciamento de itens"""
    def __init__(self, parent, main_app):
        self.parent = parent
        self.main_app = main_app
        self.frame = ttk.Frame(parent)
        self.create_widgets()
    
    def create_widgets(self):
        """Cria os widgets da view de itens"""
        ttk.Label(self.frame, text="Gerenciamento de Itens", font=('Arial', 14, 'bold')).pack(pady=10)

        # Área de formulário
        form = ttk.LabelFrame(self.frame, text="Cadastro / Edição", padding=10)
        form.pack(fill=BOTH, padx=10, pady=5)

        form_grid = ttk.Frame(form)
        form_grid.pack(fill=X)

        ttk.Label(form_grid, text="Descrição").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.desc_var = tk.StringVar()
        ttk.Entry(form_grid, textvariable=self.desc_var, width=50).grid(row=0, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(form_grid, text="Código").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.code_var = tk.StringVar()
        ttk.Entry(form_grid, textvariable=self.code_var, width=20).grid(row=0, column=3, sticky='w', padx=5, pady=2)

        ttk.Label(form_grid, text="Marca").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.brand_var = tk.StringVar()
        ttk.Entry(form_grid, textvariable=self.brand_var, width=30).grid(row=1, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(form_grid, text="Quantidade").grid(row=1, column=2, sticky='w', padx=5, pady=2)
        self.qty_var = tk.StringVar()
        ttk.Entry(form_grid, textvariable=self.qty_var, width=10).grid(row=1, column=3, sticky='w', padx=5, pady=2)

        ttk.Label(form_grid, text="Fornecedor").grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.supplier_var = tk.StringVar()
        ttk.Combobox(form_grid, textvariable=self.supplier_var, values=self.main_app.suppliers, width=30).grid(row=2, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(form_grid, text="Preço (R$)").grid(row=2, column=2, sticky='w', padx=5, pady=2)
        self.price_var = tk.StringVar()
        ttk.Entry(form_grid, textvariable=self.price_var, width=12).grid(row=2, column=3, sticky='w', padx=5, pady=2)

        # Botões de ação do formulário
        actions = ttk.Frame(form)
        actions.pack(fill=X, pady=5)
        ttk.Button(actions, text="Salvar", command=self.save_item, bootstyle=SUCCESS).pack(side=LEFT, padx=5)
        ttk.Button(actions, text="Limpar", command=self.clear_form, bootstyle=SECONDARY).pack(side=LEFT, padx=5)

        # Lista de itens
        list_frame = ttk.LabelFrame(self.frame, text="Itens", padding=10)
        list_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)

        cols = ("id", "description", "code", "brand", "status", "quantity")
        self.items_tree = ttk.Treeview(list_frame, columns=cols, show='headings')
        for col, title in zip(cols, ["ID", "Descrição", "Código", "Marca", "Status", "Qtd"]):
            self.items_tree.heading(col, text=title)
        self.items_tree.column("id", width=50, anchor='center')
        self.items_tree.column("description", width=350)
        self.items_tree.column("code", width=100)
        self.items_tree.column("brand", width=140)
        self.items_tree.column("status", width=140)
        self.items_tree.column("quantity", width=80, anchor='e')
        self.items_tree.pack(fill=BOTH, expand=True)

        # Botões de CRUD
        list_actions = ttk.Frame(list_frame)
        list_actions.pack(fill=X, pady=5)
        ttk.Button(list_actions, text="Novo", command=self.clear_form, bootstyle=SECONDARY).pack(side=LEFT, padx=5)
        ttk.Button(list_actions, text="Editar", command=self.load_selected_item, bootstyle=INFO).pack(side=LEFT, padx=5)
        ttk.Button(list_actions, text="Excluir", command=self.delete_selected_item, bootstyle=DANGER).pack(side=LEFT, padx=5)
        ttk.Button(list_actions, text="Marcar Comprado", command=lambda: self.update_status_selected("Comprado"), bootstyle=PRIMARY).pack(side=LEFT, padx=5)
        ttk.Button(list_actions, text="Marcar Parcial", command=lambda: self.update_status_selected("Parcialmente Comprado"), bootstyle=WARNING).pack(side=LEFT, padx=5)

        self.editing_id = None
        self.refresh_items()
    
    def pack(self, **kwargs):
        self.frame.pack(**kwargs)
    
    def destroy(self):
        self.frame.destroy()

    def refresh_items(self):
        for i in self.items_tree.get_children():
            self.items_tree.delete(i)
        items = ItemRepository.get_all()
        for item in items:
            self.items_tree.insert('', 'end', values=(item.id, item.description, item.code, item.brand, item.status, f"{item.quantity:.2f}"))

    def clear_form(self):
        self.editing_id = None
        self.desc_var.set("")
        self.code_var.set("")
        self.brand_var.set("")
        self.qty_var.set("")
        self.supplier_var.set("")
        self.price_var.set("")

    def load_selected_item(self):
        sel = self.items_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um item")
            return
        vals = self.items_tree.item(sel[0], 'values')
        item = ItemRepository.get_by_id(int(vals[0]))
        if not item:
            messagebox.showerror("Erro", "Item não encontrado")
            return
        self.editing_id = item.id
        self.desc_var.set(item.description)
        self.code_var.set(item.code)
        self.brand_var.set(item.brand)
        self.qty_var.set(str(item.quantity))
        # Não conseguimos inferir fornecedor/preço específicos; deixar em branco para manter
        self.supplier_var.set("")
        self.price_var.set("")

    def save_item(self):
        try:
            description = self.desc_var.get().strip()
            code = self.code_var.get().strip()
            brand = self.brand_var.get().strip() or "N/A"
            qty = float(self.qty_var.get().replace(',', '.')) if self.qty_var.get().strip() else 0.0
            supplier = self.supplier_var.get().strip()
            price = self.price_var.get().strip()
            price_val = float(price.replace('R$', '').replace(',', '.')) if price else None

            # Validações básicas
            if not description or not code or qty <= 0:
                messagebox.showwarning("Aviso", "Preencha descrição, código e quantidade > 0")
                return

            suppliers_prices = {}
            if supplier and price_val is not None:
                suppliers_prices[supplier] = price_val

            if self.editing_id:
                item = ItemRepository.get_by_id(self.editing_id)
                if not item:
                    messagebox.showerror("Erro", "Item não encontrado")
                    return
                item.description = description
                item.code = code
                item.brand = brand
                item.quantity = qty
                # mesclar preços
                if suppliers_prices:
                    item.suppliers_prices.update(suppliers_prices)
                if ItemRepository.update(item):
                    messagebox.showinfo("Sucesso", "Item atualizado")
                else:
                    messagebox.showerror("Erro", "Falha ao atualizar item")
            else:
                new_item = Item(description=description, code=code, brand=brand, status="A Comprar", quantity=qty, suppliers_prices=suppliers_prices)
                item_id = ItemRepository.create(new_item)
                if item_id:
                    messagebox.showinfo("Sucesso", "Item criado")
                else:
                    messagebox.showerror("Erro", "Falha ao criar item")

            self.refresh_items()
            self.clear_form()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar item: {e}")

    def delete_selected_item(self):
        sel = self.items_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um item")
            return
        vals = self.items_tree.item(sel[0], 'values')
        if messagebox.askyesno("Confirmar", f"Excluir item ID {vals[0]}?"):
            if ItemRepository.delete(int(vals[0])):
                messagebox.showinfo("Sucesso", "Item excluído")
                self.refresh_items()
            else:
                messagebox.showerror("Erro", "Falha ao excluir item")

    def update_status_selected(self, new_status):
        sel = self.items_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um item")
            return
        vals = self.items_tree.item(sel[0], 'values')
        item = ItemRepository.get_by_id(int(vals[0]))
        if not item:
            messagebox.showerror("Erro", "Item não encontrado")
            return
        item.status = new_status
        if ItemRepository.update(item):
            self.refresh_items()
            messagebox.showinfo("Sucesso", f"Status atualizado para {new_status}")
        else:
            messagebox.showerror("Erro", "Falha ao atualizar status")


class CompaniesView:
    """View para gerenciamento de empresas"""
    def __init__(self, parent, main_app):
        self.parent = parent
        self.main_app = main_app
        self.frame = ttk.Frame(parent)
        self.create_widgets()
    
    def create_widgets(self):
        """Cria os widgets da view de empresas"""
        ttk.Label(self.frame, text="Gerenciamento de Empresas", font=('Arial', 14, 'bold')).pack(pady=10)

        form = ttk.LabelFrame(self.frame, text="Cadastro / Edição", padding=10)
        form.pack(fill=BOTH, padx=10, pady=5)

        grid = ttk.Frame(form)
        grid.pack(fill=X)

        ttk.Label(grid, text="Nome").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.c_name = tk.StringVar()
        ttk.Entry(grid, textvariable=self.c_name, width=50).grid(row=0, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(grid, text="CNPJ").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.c_cnpj = tk.StringVar()
        ttk.Entry(grid, textvariable=self.c_cnpj, width=30).grid(row=1, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(grid, text="Comprador").grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.c_buyer = tk.StringVar()
        ttk.Entry(grid, textvariable=self.c_buyer, width=30).grid(row=2, column=1, sticky='w', padx=5, pady=2)

        actions = ttk.Frame(form)
        actions.pack(fill=X, pady=5)
        ttk.Button(actions, text="Salvar", command=self.save_company, bootstyle=SUCCESS).pack(side=LEFT, padx=5)
        ttk.Button(actions, text="Novo", command=self.clear_form, bootstyle=SECONDARY).pack(side=LEFT, padx=5)

        list_frame = ttk.LabelFrame(self.frame, text="Empresas", padding=10)
        list_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)

        cols = ("id", "name", "cnpj", "buyer")
        self.company_tree = ttk.Treeview(list_frame, columns=cols, show='headings')
        for col, title in zip(cols, ["ID", "Nome", "CNPJ", "Comprador"]):
            self.company_tree.heading(col, text=title)
        self.company_tree.column("id", width=50, anchor='center')
        self.company_tree.column("name", width=300)
        self.company_tree.column("cnpj", width=140)
        self.company_tree.column("buyer", width=180)
        self.company_tree.pack(fill=BOTH, expand=True)

        la = ttk.Frame(list_frame)
        la.pack(fill=X, pady=5)
        ttk.Button(la, text="Editar", command=self.load_selected, bootstyle=INFO).pack(side=LEFT, padx=5)
        ttk.Button(la, text="Excluir", command=self.delete_selected, bootstyle=DANGER).pack(side=LEFT, padx=5)

        self.editing_id = None
        self.refresh_companies()
    
    def pack(self, **kwargs):
        self.frame.pack(**kwargs)
    
    def destroy(self):
        self.frame.destroy()

    def refresh_companies(self):
        for i in self.company_tree.get_children():
            self.company_tree.delete(i)
        from models import CompanyRepository
        for c in CompanyRepository.get_all():
            self.company_tree.insert('', 'end', values=(c.id, c.name, c.cnpj, c.buyer_name))

    def clear_form(self):
        self.editing_id = None
        self.c_name.set("")
        self.c_cnpj.set("")
        self.c_buyer.set("")

    def load_selected(self):
        sel = self.company_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione uma empresa")
            return
        vals = self.company_tree.item(sel[0], 'values')
        from models import CompanyRepository
        company = CompanyRepository.get_by_id(int(vals[0]))
        if not company:
            messagebox.showerror("Erro", "Empresa não encontrada")
            return
        self.editing_id = company.id
        self.c_name.set(company.name)
        self.c_cnpj.set(company.cnpj)
        self.c_buyer.set(company.buyer_name)

    def save_company(self):
        from models import Company, CompanyRepository
        name = self.c_name.get().strip()
        cnpj = self.c_cnpj.get().strip()
        buyer = self.c_buyer.get().strip()
        if not name or not cnpj or not buyer:
            messagebox.showwarning("Aviso", "Preencha todos os campos")
            return
        if self.editing_id:
            company = CompanyRepository.get_by_id(self.editing_id)
            if not company:
                messagebox.showerror("Erro", "Empresa não encontrada")
                return
            company.name = name
            company.cnpj = cnpj
            company.buyer_name = buyer
            if CompanyRepository.update(company):
                messagebox.showinfo("Sucesso", "Empresa atualizada")
            else:
                messagebox.showerror("Erro", "Falha ao atualizar empresa")
        else:
            company = Company(name=name, cnpj=cnpj, buyer_name=buyer)
            if CompanyRepository.create(company):
                messagebox.showinfo("Sucesso", "Empresa criada")
            else:
                messagebox.showerror("Erro", "Falha ao criar empresa")
        self.refresh_companies()
        self.clear_form()

    def delete_selected(self):
        sel = self.company_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione uma empresa")
            return
        vals = self.company_tree.item(sel[0], 'values')
        from models import CompanyRepository
        if messagebox.askyesno("Confirmar", f"Excluir empresa ID {vals[0]}?"):
            if CompanyRepository.delete(int(vals[0])):
                messagebox.showinfo("Sucesso", "Empresa excluída")
                self.refresh_companies()
            else:
                messagebox.showerror("Erro", "Falha ao excluir empresa")


class SuppliersView:
    """View para gerenciamento de fornecedores"""
    def __init__(self, parent, main_app):
        self.parent = parent
        self.main_app = main_app
        self.frame = ttk.Frame(parent)
        self.create_widgets()
    
    def create_widgets(self):
        """Cria os widgets da view de fornecedores"""
        ttk.Label(self.frame, text="Gerenciamento de Fornecedores", font=('Arial', 14, 'bold')).pack(pady=10)

        form = ttk.LabelFrame(self.frame, text="Cadastro / Edição", padding=10)
        form.pack(fill=BOTH, padx=10, pady=5)

        grid = ttk.Frame(form)
        grid.pack(fill=X)

        ttk.Label(grid, text="Nome").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.s_name = tk.StringVar()
        ttk.Entry(grid, textvariable=self.s_name, width=40).grid(row=0, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(grid, text="CNPJ").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.s_cnpj = tk.StringVar()
        ttk.Entry(grid, textvariable=self.s_cnpj, width=30).grid(row=1, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(grid, text="Vendedor").grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.s_seller = tk.StringVar()
        ttk.Entry(grid, textvariable=self.s_seller, width=30).grid(row=2, column=1, sticky='w', padx=5, pady=2)

        actions = ttk.Frame(form)
        actions.pack(fill=X, pady=5)
        ttk.Button(actions, text="Salvar", command=self.save_supplier, bootstyle=SUCCESS).pack(side=LEFT, padx=5)
        ttk.Button(actions, text="Novo", command=self.clear_form, bootstyle=SECONDARY).pack(side=LEFT, padx=5)

        list_frame = ttk.LabelFrame(self.frame, text="Fornecedores", padding=10)
        list_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)

        cols = ("id", "name", "cnpj", "seller")
        self.supplier_tree = ttk.Treeview(list_frame, columns=cols, show='headings')
        for col, title in zip(cols, ["ID", "Nome", "CNPJ", "Vendedor"]):
            self.supplier_tree.heading(col, text=title)
        self.supplier_tree.column("id", width=50, anchor='center')
        self.supplier_tree.column("name", width=280)
        self.supplier_tree.column("cnpj", width=140)
        self.supplier_tree.column("seller", width=200)
        self.supplier_tree.pack(fill=BOTH, expand=True)

        la = ttk.Frame(list_frame)
        la.pack(fill=X, pady=5)
        ttk.Button(la, text="Editar", command=self.load_selected, bootstyle=INFO).pack(side=LEFT, padx=5)
        ttk.Button(la, text="Excluir", command=self.delete_selected, bootstyle=DANGER).pack(side=LEFT, padx=5)

        self.editing_id = None
        self.refresh_suppliers()
    
    def pack(self, **kwargs):
        self.frame.pack(**kwargs)
    
    def destroy(self):
        self.frame.destroy()

    def refresh_suppliers(self):
        for i in self.supplier_tree.get_children():
            self.supplier_tree.delete(i)
        from models import SupplierRepository
        for s in SupplierRepository.get_all():
            self.supplier_tree.insert('', 'end', values=(s.id, s.name, s.cnpj, s.seller_name))
        # atualizar lista para outras partes
        self.main_app.load_suppliers()

    def clear_form(self):
        self.editing_id = None
        self.s_name.set("")
        self.s_cnpj.set("")
        self.s_seller.set("")

    def load_selected(self):
        sel = self.supplier_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um fornecedor")
            return
        vals = self.supplier_tree.item(sel[0], 'values')
        from models import SupplierRepository
        supplier = SupplierRepository.get_by_id(int(vals[0]))
        if not supplier:
            messagebox.showerror("Erro", "Fornecedor não encontrado")
            return
        self.editing_id = supplier.id
        self.s_name.set(supplier.name)
        self.s_cnpj.set(supplier.cnpj)
        self.s_seller.set(supplier.seller_name)

    def save_supplier(self):
        from models import Supplier, SupplierRepository
        name = self.s_name.get().strip()
        cnpj = self.s_cnpj.get().strip()
        seller = self.s_seller.get().strip()
        if not name or not cnpj or not seller:
            messagebox.showwarning("Aviso", "Preencha todos os campos")
            return
        if self.editing_id:
            supplier = SupplierRepository.get_by_id(self.editing_id)
            if not supplier:
                messagebox.showerror("Erro", "Fornecedor não encontrado")
                return
            supplier.name = name
            supplier.cnpj = cnpj
            supplier.seller_name = seller
            if SupplierRepository.update(supplier):
                messagebox.showinfo("Sucesso", "Fornecedor atualizado")
            else:
                messagebox.showerror("Erro", "Falha ao atualizar fornecedor")
        else:
            supplier = Supplier(name=name, cnpj=cnpj, seller_name=seller)
            if SupplierRepository.create(supplier):
                messagebox.showinfo("Sucesso", "Fornecedor criado")
            else:
                messagebox.showerror("Erro", "Falha ao criar fornecedor")
        self.refresh_suppliers()
        self.clear_form()

    def delete_selected(self):
        sel = self.supplier_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um fornecedor")
            return
        vals = self.supplier_tree.item(sel[0], 'values')
        from models import SupplierRepository
        if messagebox.askyesno("Confirmar", f"Excluir fornecedor ID {vals[0]}?"):
            if SupplierRepository.delete(int(vals[0])):
                messagebox.showinfo("Sucesso", "Fornecedor excluído")
                self.refresh_suppliers()
            else:
                messagebox.showerror("Erro", "Falha ao excluir fornecedor")

